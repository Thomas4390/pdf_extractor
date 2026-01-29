"""
Export functionality for aggregation data.

Provides CSV and Excel export capabilities for aggregated advisor data.
"""

import io
from datetime import datetime
from typing import Optional

import pandas as pd
import streamlit as st


def export_to_csv(
    df: pd.DataFrame,
    filename_prefix: str = "aggregation",
) -> bytes:
    """
    Export DataFrame to CSV format.

    Args:
        df: DataFrame to export
        filename_prefix: Prefix for the filename

    Returns:
        CSV data as bytes
    """
    if df is None or df.empty:
        return b""

    # Create a copy to avoid modifying original
    export_df = df.copy()

    # Convert to CSV
    csv_buffer = io.StringIO()
    export_df.to_csv(csv_buffer, index=False, encoding="utf-8")
    return csv_buffer.getvalue().encode("utf-8")


def export_to_excel(
    df: pd.DataFrame,
    sheet_name: str = "Aggregation",
    filename_prefix: str = "aggregation",
) -> bytes:
    """
    Export DataFrame to Excel format with formatting.

    Args:
        df: DataFrame to export
        sheet_name: Name of the Excel sheet
        filename_prefix: Prefix for the filename

    Returns:
        Excel data as bytes
    """
    if df is None or df.empty:
        return b""

    # Create a copy to avoid modifying original
    export_df = df.copy()

    # Create Excel file in memory
    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Get the workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Format header row
        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Format numeric cells
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def get_export_filename(
    prefix: str,
    period_name: str,
    extension: str,
) -> str:
    """
    Generate a standardized export filename.

    Args:
        prefix: File prefix (e.g., "aggregation")
        period_name: Name of the period (e.g., "Janvier 2026")
        extension: File extension (e.g., "csv", "xlsx")

    Returns:
        Formatted filename
    """
    # Clean period name for filename
    clean_period = period_name.replace("/", "-").replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    return f"{prefix}_{clean_period}_{timestamp}.{extension}"


def render_export_buttons(
    df: pd.DataFrame,
    period_name: str,
    key_suffix: str = "",
) -> None:
    """
    Render CSV and Excel download buttons for the given DataFrame.

    Args:
        df: DataFrame to export
        period_name: Name of the current period for filename
        key_suffix: Suffix for widget keys to ensure uniqueness
    """
    if df is None or df.empty:
        st.info("Aucune donnée à exporter.")
        return

    col1, col2 = st.columns(2)

    # CSV export
    with col1:
        csv_data = export_to_csv(df)
        csv_filename = get_export_filename("aggregation", period_name, "csv")
        st.download_button(
            label="📥 Télécharger CSV",
            data=csv_data,
            file_name=csv_filename,
            mime="text/csv",
            key=f"export_csv_{key_suffix}",
        )

    # Excel export
    with col2:
        try:
            excel_data = export_to_excel(df, sheet_name=period_name[:31])  # Excel sheet name max 31 chars
            excel_filename = get_export_filename("aggregation", period_name, "xlsx")
            st.download_button(
                label="📥 Télécharger Excel",
                data=excel_data,
                file_name=excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"export_excel_{key_suffix}",
            )
        except ImportError:
            st.caption("Export Excel non disponible (openpyxl requis)")


def render_export_section(
    combined_df: pd.DataFrame,
    period_name: str,
) -> None:
    """
    Render a complete export section with title and buttons.

    Args:
        combined_df: Combined aggregated DataFrame
        period_name: Name of the current period
    """
    st.markdown("#### 📥 Exporter les données")
    st.caption("Téléchargez les données agrégées au format CSV ou Excel.")

    if combined_df is not None and not combined_df.empty:
        # Show data summary
        st.markdown(f"**Données disponibles:** {len(combined_df)} conseillers, {len(combined_df.columns)} colonnes")

        render_export_buttons(combined_df, period_name, key_suffix="main")
    else:
        st.warning("Aucune donnée à exporter.")
